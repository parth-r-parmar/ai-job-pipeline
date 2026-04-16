"""Export job results to a formatted Excel workbook."""

import os
import logging
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.config import OUTPUT_EXCEL, MATCH_THRESHOLD

logger = logging.getLogger(__name__)

# Styles
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="E4CCFF", end_color="E4CCFF", fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single")
PRODUCT_FONT = Font(bold=True, color="006100")
SERVICE_FONT = Font(color="9C5700")

ALL_JOBS_COLUMNS = [
    ("Title", 30),
    ("Company", 25),
    ("Company Type", 14),
    ("Location", 20),
    ("Match Score", 12),
    ("Recommendation", 18),
    ("Reason", 40),
    ("Ghost Risk", 10),
    ("Salary", 18),
    ("Salary (LPA)", 12),
    ("Experience", 15),
    ("Job Type", 12),
    ("Posted Date", 15),
    ("Source", 12),
    ("Matched Keywords", 35),
    ("Missing Keywords", 35),
    ("Injected Keywords", 35),
    ("URL", 45),
    ("Tailored Resume", 30),
]

GHOST_HIGH_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GHOST_MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

COLD_OUTREACH_COLUMNS = [
    ("Company", 30),
    ("Company Type", 14),
    ("Why Reach Out", 50),
    ("Roles Seen", 35),
    ("Locations", 25),
    ("LinkedIn URL", 45),
    ("Suggested Action", 40),
]

SUMMARY_LABEL_FONT = Font(bold=True, size=11)
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")


def _score_fill(score: int) -> PatternFill:
    if score >= MATCH_THRESHOLD:
        return GREEN_FILL
    elif score >= 40:
        return YELLOW_FILL
    return RED_FILL


def _write_header(ws, columns: list[tuple[str, int]]):
    for col, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def _write_job_row(ws, row_idx: int, job: dict):
    score = job.get("match_score", 0)
    company_type = job.get("company_type", "unknown")
    salary_lpa = job.get("salary_lpa")
    ghost_risk = job.get("ghost_risk", "low")
    injected = ", ".join(job.get("core_competencies", []))

    values = [
        job.get("title", ""),                               # 1
        job.get("company", ""),                              # 2
        company_type.title(),                                # 3
        job.get("location", ""),                             # 4
        score,                                               # 5
        job.get("recommendation", ""),                       # 6
        job.get("reason", ""),                               # 7
        ghost_risk.title(),                                  # 8
        job.get("salary", ""),                               # 9
        f"{salary_lpa} LPA" if salary_lpa else "Not disclosed",  # 10
        job.get("experience", ""),                           # 11
        job.get("job_type", ""),                             # 12
        job.get("posted_date", ""),                          # 13
        job.get("source", ""),                               # 14
        ", ".join(job.get("matched_keywords", [])),          # 15
        ", ".join(job.get("missing_keywords", [])),          # 16
        injected,                                            # 17
        job.get("url", ""),                                  # 18
        job.get("tailored_pdf_path", ""),                    # 19
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.fill = _score_fill(score)
        # Company type highlight
        if col == 3:
            if company_type == "product":
                cell.font = PRODUCT_FONT
                cell.fill = BLUE_FILL
            elif company_type == "service":
                cell.font = SERVICE_FONT
        # Ghost risk coloring
        if col == 8:
            if ghost_risk == "high":
                cell.fill = GHOST_HIGH_FILL
            elif ghost_risk == "medium":
                cell.fill = GHOST_MED_FILL
        # URL clickable
        if col == 18 and val:
            cell.hyperlink = val
            cell.font = LINK_FONT


def _build_all_jobs_sheet(ws, jobs: list[dict]):
    _write_header(ws, ALL_JOBS_COLUMNS)
    # Sort: product companies first, then by score desc
    sorted_jobs = sorted(
        jobs,
        key=lambda j: (
            0 if j.get("company_type") == "product" else 1,
            -(j.get("match_score", 0)),
            -(j.get("salary_lpa") if isinstance(j.get("salary_lpa"), (int, float)) else 0),
        ),
    )
    for i, job in enumerate(sorted_jobs, 2):
        _write_job_row(ws, i, job)


def _build_top_matches_sheet(ws, jobs: list[dict]):
    _write_header(ws, ALL_JOBS_COLUMNS)
    top = sorted(
        [j for j in jobs if j.get("match_score", 0) >= MATCH_THRESHOLD],
        key=lambda j: (
            0 if j.get("company_type") == "product" else 1,
            -(j.get("match_score", 0)),
        ),
    )
    for i, job in enumerate(top, 2):
        _write_job_row(ws, i, job)


def _build_cold_outreach_sheet(ws, jobs: list[dict]):
    """Sheet for companies worth cold-emailing.

    Identifies:
    1. Product companies from scraped results (good learning environment)
    2. Companies with multiple listings (actively growing)
    3. Well-known product companies not in results (suggest proactive outreach)
    """
    _write_header(ws, COLD_OUTREACH_COLUMNS)

    outreach_rows = []

    # Group jobs by company
    company_jobs: dict[str, list[dict]] = {}
    for job in jobs:
        name = job.get("company", "").strip()
        if name and name != "N/A":
            company_jobs.setdefault(name, []).append(job)

    # 1. Product companies from results — great targets
    for company, cjobs in company_jobs.items():
        ctype = cjobs[0].get("company_type", "unknown")
        if ctype == "product":
            roles = list(set(j.get("title", "") for j in cjobs))[:5]
            locations = list(set(j.get("location", "") for j in cjobs if j.get("location")))[:3]
            outreach_rows.append({
                "company": company,
                "company_type": "Product",
                "why": f"Product company with {len(cjobs)} active listing(s). Good for growth and learning.",
                "roles": ", ".join(roles),
                "locations": ", ".join(locations),
                "linkedin": f"https://www.linkedin.com/company/{company.lower().replace(' ', '-')}/jobs/",
                "action": "Apply directly + send cold LinkedIn message to engineering manager",
            })

    # 2. Companies with 3+ listings (actively hiring, growing team)
    for company, cjobs in company_jobs.items():
        ctype = cjobs[0].get("company_type", "unknown")
        if ctype != "product" and len(cjobs) >= 3:
            roles = list(set(j.get("title", "") for j in cjobs))[:5]
            locations = list(set(j.get("location", "") for j in cjobs if j.get("location")))[:3]
            outreach_rows.append({
                "company": company,
                "company_type": ctype.title(),
                "why": f"Actively hiring ({len(cjobs)} listings). Growing team = more opportunities.",
                "roles": ", ".join(roles),
                "locations": ", ".join(locations),
                "linkedin": f"https://www.linkedin.com/company/{company.lower().replace(' ', '-')}/jobs/",
                "action": "Apply to best-fit role + cold email HR/recruiter with tailored resume",
            })

    # 3. Well-known product companies to proactively reach out to
    # (not in current results but worth cold-emailing)
    from src.filters import KNOWN_PRODUCT_COMPANIES
    seen_companies = {c.lower().strip() for c in company_jobs.keys()}

    # Curated list of top product companies in India with good engineering culture
    recommended_companies = [
        ("Razorpay", "Fintech leader. Strong engineering culture, great learning."),
        ("Zerodha", "Bootstrapped, profitable. Small team = high ownership."),
        ("CRED", "Premium fintech. Invests heavily in engineering excellence."),
        ("Postman", "API platform used by millions. Remote-friendly, strong JS/TS stack."),
        ("BrowserStack", "Testing platform. Remote-first, uses React/Node heavily."),
        ("Freshworks", "SaaS unicorn. Chennai + India offices. Good work-life balance."),
        ("Meesho", "Social commerce. Fast growth, modern tech stack."),
        ("Groww", "Investment platform. Bangalore-based, strong engineering."),
        ("PhonePe", "Digital payments. Large scale, good compensation."),
        ("Juspay", "Payment orchestration. Functional programming, great engineering."),
        ("Slice", "Fintech. Modern stack, startup culture."),
        ("Jupiter", "Neobank. Small team, fast-paced, full-stack roles."),
        ("ShareChat", "Social media. Uses React, good for frontend roles."),
        ("Flipkart", "E-commerce giant. Great engineering practices and mentorship."),
        ("Swiggy", "Food delivery + quick commerce. Large-scale React/Node systems."),
        ("CleverTap", "Martech SaaS. Mumbai-based, uses React."),
        ("MoEngage", "Customer engagement platform. Good for full-stack JS devs."),
        ("Rocketlane", "Customer onboarding SaaS. Small team, high impact."),
        ("Toplyne", "Product-led growth platform. Modern tech, early stage."),
        ("Vercel", "Next.js creators. Remote-first, dream for React developers."),
    ]

    for company_name, reason in recommended_companies:
        if company_name.lower() not in seen_companies:
            outreach_rows.append({
                "company": company_name,
                "company_type": "Product",
                "why": reason,
                "roles": "Full Stack / React Developer",
                "locations": "Check careers page",
                "linkedin": f"https://www.linkedin.com/company/{company_name.lower().replace(' ', '-')}/jobs/",
                "action": "Check careers page. If no open role, cold email with portfolio + tailored resume.",
            })

    # Write rows
    for i, row in enumerate(outreach_rows, 2):
        ws.cell(row=i, column=1, value=row["company"])
        ws.cell(row=i, column=2, value=row["company_type"])
        ws.cell(row=i, column=3, value=row["why"])
        ws.cell(row=i, column=4, value=row["roles"])
        ws.cell(row=i, column=5, value=row["locations"])

        link_cell = ws.cell(row=i, column=6, value=row["linkedin"])
        link_cell.hyperlink = row["linkedin"]
        link_cell.font = LINK_FONT

        ws.cell(row=i, column=7, value=row["action"])

        # Highlight product companies
        if row["company_type"] == "Product":
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = BLUE_FILL


def _build_summary_sheet(ws, jobs: list[dict]):
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    rows = []

    # Total
    rows.append(("Total Jobs (after filters)", len(jobs), True))
    rows.append(("", "", False))

    # Per source
    rows.append(("Jobs by Source", "", True))
    source_counts = Counter(j.get("source", "unknown") for j in jobs)
    for source, count in sorted(source_counts.items()):
        rows.append((f"  {source}", count, False))
    rows.append(("", "", False))

    # Company type breakdown
    rows.append(("Company Type Breakdown", "", True))
    type_counts = Counter(j.get("company_type", "unknown") for j in jobs)
    for ctype, count in type_counts.most_common():
        rows.append((f"  {ctype.title()}", count, False))
    rows.append(("", "", False))

    # Salary stats
    raw_salaries = [j["salary_lpa"] for j in jobs if j.get("salary_lpa")]
    salaries = []
    for s in raw_salaries:
        if isinstance(s, (int, float)):
            salaries.append(s)
        elif isinstance(s, str):
            try:
                salaries.append(float(s.replace(" LPA", "").replace(",", "")))
            except ValueError:
                pass
    if salaries:
        rows.append(("Salary Statistics", "", True))
        rows.append((f"  Jobs with salary disclosed", len(salaries), False))
        rows.append((f"  Minimum salary", f"{min(salaries)} LPA", False))
        rows.append((f"  Maximum salary", f"{max(salaries)} LPA", False))
        rows.append((f"  Average salary", f"{round(sum(salaries)/len(salaries), 1)} LPA", False))
        rows.append(("", "", False))

    # Remote vs on-site
    remote_count = sum(1 for j in jobs if "remote" in j.get("location", "").lower())
    rows.append(("Location Breakdown", "", True))
    rows.append((f"  Remote / Remote-friendly", remote_count, False))
    rows.append((f"  On-site / Hybrid", len(jobs) - remote_count, False))
    rows.append(("", "", False))

    # Score distribution
    rows.append(("Score Distribution", "", True))
    scores = [j.get("match_score", 0) for j in jobs]
    if scores and any(s > 0 for s in scores):
        rows.append(("  Strong Match (90-100)", sum(1 for s in scores if s >= 90), False))
        rows.append(("  Good Match (70-89)", sum(1 for s in scores if 70 <= s < 90), False))
        rows.append(("  Moderate Match (40-69)", sum(1 for s in scores if 40 <= s < 70), False))
        rows.append(("  Weak/Poor Match (0-39)", sum(1 for s in scores if 0 < s < 40), False))
        rows.append(("", "", False))
        rows.append(("Average Match Score", round(sum(scores) / len(scores), 1), True))
    else:
        rows.append(("  (scoring not run — set ANTHROPIC_API_KEY)", "", False))
    rows.append(("", "", False))

    # Tailoring stats
    tailored_count = sum(1 for j in jobs if j.get("tailored_pdf_path"))
    rows.append(("Tailored Resumes Generated", tailored_count, True))

    # Top matched keywords
    all_matched = []
    for j in jobs:
        all_matched.extend(j.get("matched_keywords", []))
    if all_matched:
        rows.append(("", "", False))
        rows.append(("Top Matched Keywords", "", True))
        for kw, count in Counter(all_matched).most_common(10):
            rows.append((f"  {kw}", count, False))

    # Top missing keywords (skills to learn!)
    all_missing = []
    for j in jobs:
        all_missing.extend(j.get("missing_keywords", []))
    if all_missing:
        rows.append(("", "", False))
        rows.append(("Skills Gap (top missing keywords)", "", True))
        for kw, count in Counter(all_missing).most_common(10):
            rows.append((f"  {kw}", count, False))

    # Write rows
    for i, (label, value, is_header) in enumerate(rows, 1):
        cell_a = ws.cell(row=i, column=1, value=label)
        cell_b = ws.cell(row=i, column=2, value=value)
        if is_header and label:
            cell_a.font = SUMMARY_LABEL_FONT


def export(jobs: list[dict], output_path: str = None) -> str:
    """Create the Excel workbook with 4 sheets.

    Returns absolute path to the generated xlsx file.
    """
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "All Jobs"
    _build_all_jobs_sheet(ws1, jobs)

    ws2 = wb.create_sheet("Top Matches")
    _build_top_matches_sheet(ws2, jobs)

    ws3 = wb.create_sheet("Cold Outreach")
    _build_cold_outreach_sheet(ws3, jobs)

    ws4 = wb.create_sheet("Summary")
    _build_summary_sheet(ws4, jobs)

    path = output_path or OUTPUT_EXCEL
    wb.save(path)
    logger.info(f"Excel exported: {os.path.abspath(path)}")
    return os.path.abspath(path)
