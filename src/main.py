"""
AI Job Pipeline CLI — Scrape jobs, score matches, tailor resumes, export results.

Usage:
    python -m src.main --keywords "react developer" --location "India" --pages 2
    python -m src.main --dry-run
    python -m src.main --demo
    python -m src.main --scorer cli --pdf-style modern
"""

import argparse
import json
import logging
import os
import time
import webbrowser

from src.config import (
    MATCH_THRESHOLD, PAGES_PER_SITE, OUTPUT_EXCEL,
    ANTHROPIC_API_KEY, INCLUDE_REMOTE, OUTPUT_DIR,
    MAX_DETAIL_FETCHES, PDF_STYLE, PROJECT_ROOT,
)
from src.resume_parser import get_search_keywords
from src.scrapers import ALL_SCRAPERS
from src.filters import filter_jobs
from src.ghost_detector import flag_ghost_jobs
from src.scorer import score_all_jobs
from src.tailor import tailor_and_generate
from src.exporter import export

logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
)
logger = logging.getLogger(__name__)

_SCRAPER_MAP = {s.SOURCE: s for s in ALL_SCRAPERS}

TAILORED_DIR = str(OUTPUT_DIR / "tailored_resumes")


# ---------------------------------------------------------------------------
# Phase header helpers
# ---------------------------------------------------------------------------

def _phase(title: str):
    logger.info(f"\n{'═' * 3} {title} {'═' * (60 - len(title))}")


def _done(start_time: float, jobs: list[dict]):
    elapsed = time.time() - start_time
    mins = int(elapsed // 60)
    secs = int(elapsed % 60)
    tailored = sum(1 for j in jobs if j.get("tailored_pdf_path"))
    logger.info(f"\n{'═' * 3} DONE ({mins}m {secs}s) {'═' * 47}")
    logger.info(f"  Excel: {OUTPUT_EXCEL}")
    logger.info(f"  Jobs: {len(jobs)} total, {sum(1 for j in jobs if j.get('match_score', 0) >= MATCH_THRESHOLD)} matched, {tailored} tailored")
    if tailored:
        logger.info(f"  PDFs:  {TAILORED_DIR}/")


# ---------------------------------------------------------------------------
# Cross-run dedup: load previous results
# ---------------------------------------------------------------------------

def _load_previous_results(excel_path: str) -> dict:
    """Load previously scored jobs from existing Excel. Returns {url: score_fields}."""
    if not os.path.exists(excel_path):
        return {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel_path, read_only=True)
        if "All Jobs" not in wb.sheetnames:
            return {}
        ws = wb["All Jobs"]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        prev = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            job = dict(zip(headers, row))
            url = job.get("URL") or f"{job.get('Title', '')}|{job.get('Company', '')}"
            score = job.get("Match Score")
            if url and score and score > 0:
                prev[url] = {
                    "match_score": score,
                    "recommendation": job.get("Recommendation", ""),
                    "reason": job.get("Reason", ""),
                    "matched_keywords": (job.get("Matched Keywords") or "").split(", "),
                    "missing_keywords": (job.get("Missing Keywords") or "").split(", "),
                }
        wb.close()
        return prev
    except Exception:
        return {}


def _job_key(job: dict) -> str:
    return job.get("url") or f"{job.get('title', '')}|{job.get('company', '')}"


def _tailor_exists(job: dict) -> str | None:
    """Check if tailored JSON already exists. Returns path if found."""
    import re
    company = re.sub(r"[^\w\-]", "", job.get("company", "unknown").replace(" ", "_"))[:50]
    title = re.sub(r"[^\w\-]", "", job.get("title", "unknown").replace(" ", "_"))[:50]
    json_path = os.path.join(TAILORED_DIR, f"{company}_{title}.json")
    pdf_path = os.path.join(TAILORED_DIR, f"{company}_{title}.pdf")
    if os.path.exists(json_path) and os.path.exists(pdf_path):
        return os.path.abspath(pdf_path)
    return None


# ---------------------------------------------------------------------------
# CLI args
# ---------------------------------------------------------------------------

def parse_args():
    parser = argparse.ArgumentParser(
        description="AI-powered job search pipeline — scrape, score, tailor, export."
    )
    parser.add_argument(
        "--keywords", "-k", nargs="+",
        help="Search keywords (default: auto-extracted from resume)",
    )
    parser.add_argument(
        "--location", "-l", default="India",
        help="Job location filter (default: India)",
    )
    parser.add_argument(
        "--pages", "-p", type=int, default=PAGES_PER_SITE,
        help=f"Pages to scrape per site (default: {PAGES_PER_SITE})",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Scrape + score only, skip resume tailoring (saves API cost)",
    )
    parser.add_argument(
        "--output", "-o", default=OUTPUT_EXCEL,
        help=f"Output Excel path (default: {OUTPUT_EXCEL})",
    )
    parser.add_argument(
        "--scrapers", "-s", nargs="+", choices=list(_SCRAPER_MAP.keys()),
        help="Run only specific scrapers (default: all)",
    )
    parser.add_argument(
        "--no-remote", action="store_true",
        help="Skip the remote jobs search pass",
    )
    parser.add_argument(
        "--pdf-style", choices=["classic", "modern", "both"], default=PDF_STYLE,
        help="Resume PDF style: classic (fpdf2), modern (HTML+Playwright), or both",
    )
    parser.add_argument(
        "--scorer", choices=["api", "cli"], default="api",
        help="AI backend: api (needs ANTHROPIC_API_KEY) or cli (Claude Code CLI)",
    )
    parser.add_argument(
        "--open-top", type=int, default=0,
        help="Open top N job URLs in browser after export",
    )
    parser.add_argument(
        "--schedule", choices=["daily", "weekly", "off"],
        help="Set up recurring scan: daily (9AM), weekly (Monday 9AM), or off",
    )
    parser.add_argument(
        "--demo", action="store_true",
        help="Run with sample data (no scraping, no AI). Shows what the output looks like.",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Ignore cached scores and tailored resumes. Re-process everything from scratch.",
    )
    return parser.parse_args()


# ---------------------------------------------------------------------------
# Scraping helpers
# ---------------------------------------------------------------------------

def _scrape_pass(scraper_classes, keywords, location, pages, label=""):
    jobs = []
    tag = f" ({label})" if label else ""
    for ScraperClass in scraper_classes:
        scraper = ScraperClass()
        try:
            result = scraper.scrape(keywords, location, pages)
            logger.info(f"  [{scraper.SOURCE}] {len(result)} jobs{tag}")
            jobs.extend(result)
        except Exception as e:
            logger.error(f"  [{scraper.SOURCE}] FAILED: {e}")
    return jobs


def _deduplicate(jobs: list[dict]) -> list[dict]:
    seen = set()
    unique = []
    for job in jobs:
        key = _job_key(job)
        if key and key not in seen:
            seen.add(key)
            unique.append(job)
    removed = len(jobs) - len(unique)
    if removed:
        logger.info(f"  Removed {removed} duplicates")
    return unique


def _open_top_jobs(jobs: list[dict], count: int):
    scored = sorted(jobs, key=lambda j: j.get("match_score", 0), reverse=True)
    opened = 0
    for job in scored:
        if opened >= count:
            break
        url = job.get("url")
        if url:
            webbrowser.open(url)
            opened += 1
            time.sleep(0.5)
    logger.info(f"  Opened {opened} job URLs in browser")


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main():
    args = parse_args()
    start_time = time.time()

    # Handle scheduling
    if args.schedule:
        from src.scheduler import setup_schedule
        setup_schedule(args.schedule, args)
        return

    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ---- Demo mode ----
    if args.demo:
        _phase("DEMO MODE (sample data, no scraping, no AI)")
        sample_path = PROJECT_ROOT / "examples" / "sample_jobs.json"
        if not sample_path.exists():
            logger.error("Sample data not found. Expected: examples/sample_jobs.json")
            return
        with open(sample_path, "r", encoding="utf-8") as f:
            all_jobs = json.load(f)
        logger.info(f"  Loaded {len(all_jobs)} sample jobs")

        _phase("GHOST DETECTION (Python heuristics)")
        flag_ghost_jobs(all_jobs)

        _phase("FILTERING")
        all_jobs = filter_jobs(all_jobs)
        logger.info(f"  {len(all_jobs)} jobs after filters")

        _phase("EXPORTING")
        output = export(all_jobs, args.output)
        logger.info(f"  Excel: {output}")

        _done(start_time, all_jobs)
        return

    # ---- Live pipeline ----

    # Step 1: Keywords
    keywords = args.keywords or get_search_keywords()
    scraper_names = ", ".join(args.scrapers) if args.scrapers else "all"
    _phase(f"SCRAPING ({scraper_names}, {args.pages} pages, location: {args.location})")

    scraper_classes = (
        [_SCRAPER_MAP[name] for name in args.scrapers]
        if args.scrapers else ALL_SCRAPERS
    )

    all_jobs = _scrape_pass(scraper_classes, keywords, args.location, args.pages)

    # Remote pass
    if INCLUDE_REMOTE and not args.no_remote:
        logger.info("  --- Remote pass ---")
        remote_jobs = _scrape_pass(
            scraper_classes, keywords, "Remote", args.pages, label="remote"
        )
        for job in remote_jobs:
            if job.get("location") in ("N/A", "", None):
                job["location"] = "Remote"
            elif "remote" not in job.get("location", "").lower():
                job["location"] = job["location"] + " (Remote)"
        all_jobs.extend(remote_jobs)

    # Deduplicate
    all_jobs = _deduplicate(all_jobs)
    logger.info(f"  Total unique: {len(all_jobs)}")

    # Enrich descriptions
    _phase("ENRICHING (fetching full job descriptions)")
    for ScraperClass in scraper_classes:
        scraper = ScraperClass()
        source_jobs = [j for j in all_jobs if j.get("source") == scraper.SOURCE]
        scraper.enrich_descriptions(source_jobs, max_fetches=MAX_DETAIL_FETCHES)

    # Ghost detection
    _phase("GHOST DETECTION (Python heuristics)")
    flag_ghost_jobs(all_jobs)

    # Filter
    _phase("FILTERING (salary, internships, company type)")
    all_jobs = filter_jobs(all_jobs)
    if not all_jobs:
        logger.warning("  No jobs remaining after filters. Exiting.")
        return
    logger.info(f"  {len(all_jobs)} jobs after filters")

    # Score
    if args.scorer == "cli":
        from src.scorer_cli import score_all_jobs as _score_fn
        from src.tailor_cli import tailor_and_generate as _tailor_fn
        ai_available = True
        ai_label = "Claude CLI"
    else:
        _score_fn = score_all_jobs
        _tailor_fn = tailor_and_generate
        ai_available = bool(ANTHROPIC_API_KEY)
        ai_label = "Anthropic API"

    if ai_available:
        _phase(f"SCORING ({len(all_jobs)} jobs via {ai_label})")

        # Load previous results for dedup (unless --force)
        prev = {} if args.force else _load_previous_results(args.output)
        skipped_score = 0

        for i, job in enumerate(all_jobs, 1):
            key = _job_key(job)
            if key in prev:
                job.update(prev[key])
                skipped_score += 1
                logger.info(f"  [{i}/{len(all_jobs)}] SKIP (cached): {job.get('title')} at {job.get('company')}")
            else:
                logger.info(f"  [{i}/{len(all_jobs)}] {job.get('title')} at {job.get('company')}")
                result = _score_fn([job])  # score single job
                if i < len(all_jobs):
                    time.sleep(0.5)

        if skipped_score:
            logger.info(f"  Reused {skipped_score} cached scores, scored {len(all_jobs) - skipped_score} new")

        scored_above = sum(1 for j in all_jobs if j.get("match_score", 0) >= MATCH_THRESHOLD)
        logger.info(f"  {scored_above} jobs scored >= {MATCH_THRESHOLD}")

        # Tailor (unless dry-run)
        if not args.dry_run:
            top_jobs = [j for j in all_jobs if j.get("match_score", 0) >= MATCH_THRESHOLD]
            if top_jobs:
                _phase(f"TAILORING ({len(top_jobs)} top matches, {args.pdf_style} PDF)")
                os.makedirs(TAILORED_DIR, exist_ok=True)
                skipped_tailor = 0

                for i, job in enumerate(top_jobs, 1):
                    # Check if already tailored (unless --force)
                    if not args.force:
                        existing = _tailor_exists(job)
                        if existing:
                            job["tailored_pdf_path"] = existing
                            skipped_tailor += 1
                            logger.info(f"  [{i}/{len(top_jobs)}] SKIP (cached): {job.get('company')}")
                            continue

                    logger.info(f"  [{i}/{len(top_jobs)}] {job.get('title')} at {job.get('company')}")
                    pdf_path = _tailor_fn(job, pdf_style=args.pdf_style)
                    if pdf_path:
                        job["tailored_pdf_path"] = str(pdf_path) if not isinstance(pdf_path, list) else " | ".join(pdf_path)
                    time.sleep(0.5)

                if skipped_tailor:
                    logger.info(f"  Reused {skipped_tailor} cached, tailored {len(top_jobs) - skipped_tailor} new")
            else:
                logger.info("  No jobs above threshold — skipping tailoring.")
        else:
            logger.info("  Dry run — skipping resume tailoring.")
    else:
        logger.warning(
            "  No AI backend available. Use --scorer cli or set ANTHROPIC_API_KEY in .env"
        )

    # Export
    _phase("EXPORTING")
    output = export(all_jobs, args.output)
    logger.info(f"  Excel: {output}")

    _done(start_time, all_jobs)

    # Open top jobs in browser
    if args.open_top > 0:
        _open_top_jobs(all_jobs, args.open_top)


if __name__ == "__main__":
    main()
