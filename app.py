"""
AI Job Pipeline — Web GUI

Run: python app.py
Opens: http://localhost:5000
"""

import json
import logging
import os
import queue
import sys
import threading
import time
import webbrowser
import zipfile
from io import BytesIO
from pathlib import Path

from flask import Flask, render_template, request, Response, jsonify, send_file, send_from_directory

# Add project root to path
sys.path.insert(0, os.path.dirname(__file__))

from src.config import (
    OUTPUT_DIR, OUTPUT_EXCEL, MATCH_THRESHOLD, PROJECT_ROOT,
    ANTHROPIC_API_KEY,
)
from src.resume_parser import get_search_keywords, get_resume
from src.filters import filter_jobs
from src.ghost_detector import flag_ghost_jobs
from src.exporter import export as export_excel

app = Flask(__name__)

TAILORED_DIR = OUTPUT_DIR / "tailored_resumes"

# Global state for SSE
_log_queue = queue.Queue()
_pipeline_running = False
_pipeline_thread = None


# ---------------------------------------------------------------------------
# Custom log handler that feeds the SSE queue
# ---------------------------------------------------------------------------

class QueueLogHandler(logging.Handler):
    def emit(self, record):
        msg = self.format(record)
        _log_queue.put({"type": "log", "message": msg})


# ---------------------------------------------------------------------------
# Pipeline runner (runs in background thread)
# ---------------------------------------------------------------------------

def _run_pipeline(params):
    global _pipeline_running
    _pipeline_running = True

    try:
        handler = QueueLogHandler()
        handler.setFormatter(logging.Formatter("%(message)s"))
        root_logger = logging.getLogger()
        root_logger.addHandler(handler)
        root_logger.setLevel(logging.INFO)

        start_time = time.time()

        # Demo mode
        if params.get("demo"):
            _log_queue.put({"type": "phase", "name": "DEMO MODE"})
            sample_path = PROJECT_ROOT / "examples" / "sample_jobs.json"
            with open(sample_path, "r", encoding="utf-8") as f:
                all_jobs = json.load(f)
            _log_queue.put({"type": "log", "message": f"Loaded {len(all_jobs)} sample jobs"})
            flag_ghost_jobs(all_jobs)
            all_jobs = filter_jobs(all_jobs)
            output = export_excel(all_jobs, str(OUTPUT_DIR / "jobs.xlsx"))
            elapsed = round(time.time() - start_time, 1)
            _log_queue.put({"type": "done", "time": f"{elapsed}s", "total": len(all_jobs),
                           "matched": sum(1 for j in all_jobs if j.get("match_score", 0) >= MATCH_THRESHOLD),
                           "tailored": 0})
            root_logger.removeHandler(handler)
            _pipeline_running = False
            return

        # Live pipeline
        from src.scrapers import ALL_SCRAPERS
        from src.main import _scrape_pass, _deduplicate, _job_key, _load_previous_results, _tailor_exists

        _SCRAPER_MAP = {s.SOURCE: s for s in ALL_SCRAPERS}

        keywords = params.get("keywords", "").split() or get_search_keywords()
        location = params.get("location", "India")
        pages = int(params.get("pages", 2))
        scraper_names = params.get("scrapers", ["linkedin", "internshala"])
        scorer_mode = params.get("scorer", "none")
        pdf_style = params.get("pdf_style", "modern")
        min_salary = float(params.get("min_salary", 8))
        threshold = int(params.get("threshold", 70))
        include_remote = params.get("include_remote", True)
        force = params.get("force", False)
        dry_run = params.get("dry_run", False)

        # Override config for this run
        from src import config
        config.MIN_SALARY_LPA = min_salary
        config.MATCH_THRESHOLD = threshold

        scraper_classes = [_SCRAPER_MAP[n] for n in scraper_names if n in _SCRAPER_MAP]

        # Scrape
        _log_queue.put({"type": "phase", "name": "SCRAPING"})
        all_jobs = _scrape_pass(scraper_classes, keywords, location, pages)

        if include_remote:
            _log_queue.put({"type": "log", "message": "Remote pass..."})
            remote_jobs = _scrape_pass(scraper_classes, keywords, "Remote", pages, label="remote")
            for job in remote_jobs:
                if job.get("location") in ("N/A", "", None):
                    job["location"] = "Remote"
                elif "remote" not in job.get("location", "").lower():
                    job["location"] = job["location"] + " (Remote)"
            all_jobs.extend(remote_jobs)

        all_jobs = _deduplicate(all_jobs)
        _log_queue.put({"type": "log", "message": f"Total unique: {len(all_jobs)}"})

        # Enrich
        _log_queue.put({"type": "phase", "name": "ENRICHING"})
        from src.config import MAX_DETAIL_FETCHES
        for ScraperClass in scraper_classes:
            scraper = ScraperClass()
            source_jobs = [j for j in all_jobs if j.get("source") == scraper.SOURCE]
            scraper.enrich_descriptions(source_jobs, max_fetches=MAX_DETAIL_FETCHES)

        # Ghost detection
        _log_queue.put({"type": "phase", "name": "GHOST DETECTION"})
        flag_ghost_jobs(all_jobs)

        # Filter
        _log_queue.put({"type": "phase", "name": "FILTERING"})
        all_jobs = filter_jobs(all_jobs)
        _log_queue.put({"type": "log", "message": f"{len(all_jobs)} jobs after filters"})

        if not all_jobs:
            _log_queue.put({"type": "done", "time": "0s", "total": 0, "matched": 0, "tailored": 0})
            root_logger.removeHandler(handler)
            _pipeline_running = False
            return

        # Score
        if scorer_mode != "none":
            _log_queue.put({"type": "phase", "name": "SCORING"})

            if scorer_mode == "cli":
                from src.scorer_cli import score_job as _score_one
                from src.tailor_cli import tailor_and_generate as _tailor_fn
            else:
                from src.scorer import score_job as _score_one
                from src.tailor import tailor_and_generate as _tailor_fn

            prev = {} if force else _load_previous_results(str(OUTPUT_DIR / "jobs.xlsx"))

            for i, job in enumerate(all_jobs, 1):
                key = _job_key(job)
                if key in prev:
                    job.update(prev[key])
                    _log_queue.put({"type": "progress", "current": i, "total": len(all_jobs),
                                   "message": f"SKIP (cached): {job.get('title')}"})
                else:
                    _log_queue.put({"type": "progress", "current": i, "total": len(all_jobs),
                                   "message": f"{job.get('title')} at {job.get('company')}"})
                    result = _score_one(job)
                    job.update(result)
                    if i < len(all_jobs):
                        time.sleep(0.3)

            # Tailor
            if not dry_run:
                top_jobs = [j for j in all_jobs if j.get("match_score", 0) >= threshold]
                if top_jobs:
                    _log_queue.put({"type": "phase", "name": "TAILORING"})
                    os.makedirs(str(TAILORED_DIR), exist_ok=True)

                    for i, job in enumerate(top_jobs, 1):
                        if not force:
                            existing = _tailor_exists(job)
                            if existing:
                                job["tailored_pdf_path"] = existing
                                _log_queue.put({"type": "progress", "current": i, "total": len(top_jobs),
                                               "message": f"SKIP (cached): {job.get('company')}"})
                                continue

                        _log_queue.put({"type": "progress", "current": i, "total": len(top_jobs),
                                       "message": f"{job.get('title')} at {job.get('company')}"})
                        pdf_path = _tailor_fn(job, pdf_style=pdf_style)
                        if pdf_path:
                            job["tailored_pdf_path"] = str(pdf_path) if not isinstance(pdf_path, list) else " | ".join(pdf_path)
                        time.sleep(0.3)

        # Export
        _log_queue.put({"type": "phase", "name": "EXPORTING"})
        output = export_excel(all_jobs, str(OUTPUT_DIR / "jobs.xlsx"))

        elapsed = round(time.time() - start_time, 1)
        mins = int(elapsed // 60)
        secs = int(elapsed % 60)
        tailored = sum(1 for j in all_jobs if j.get("tailored_pdf_path"))
        matched = sum(1 for j in all_jobs if j.get("match_score", 0) >= threshold)

        _log_queue.put({"type": "done", "time": f"{mins}m {secs}s", "total": len(all_jobs),
                       "matched": matched, "tailored": tailored})

        root_logger.removeHandler(handler)

    except Exception as e:
        _log_queue.put({"type": "error", "message": str(e)})

    _pipeline_running = False


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    auto_keywords = " ".join(get_search_keywords()[:5])
    has_results = os.path.exists(str(OUTPUT_DIR / "jobs.xlsx"))
    return render_template("index.html", auto_keywords=auto_keywords, has_results=has_results)


@app.route("/run")
def run():
    global _pipeline_thread

    if _pipeline_running:
        return jsonify({"error": "Pipeline already running"}), 409

    # Clear queue
    while not _log_queue.empty():
        _log_queue.get_nowait()

    params = {
        "keywords": request.args.get("keywords", ""),
        "location": request.args.get("location", "India"),
        "pages": request.args.get("pages", "2"),
        "scrapers": request.args.getlist("scrapers") or ["linkedin", "internshala"],
        "scorer": request.args.get("scorer", "none"),
        "pdf_style": request.args.get("pdf_style", "modern"),
        "min_salary": request.args.get("min_salary", "8"),
        "threshold": request.args.get("threshold", "70"),
        "include_remote": request.args.get("include_remote", "true") == "true",
        "force": request.args.get("force", "false") == "true",
        "dry_run": request.args.get("dry_run", "false") == "true",
        "demo": request.args.get("demo", "false") == "true",
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    _pipeline_thread = threading.Thread(target=_run_pipeline, args=(params,), daemon=True)
    _pipeline_thread.start()

    def generate():
        while True:
            try:
                msg = _log_queue.get(timeout=120)
                yield f"data: {json.dumps(msg)}\n\n"
                if msg.get("type") in ("done", "error"):
                    break
            except queue.Empty:
                yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"

    return Response(generate(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.route("/results")
def results():
    excel_path = str(OUTPUT_DIR / "jobs.xlsx")
    if not os.path.exists(excel_path):
        return jsonify([])

    from openpyxl import load_workbook
    wb = load_workbook(excel_path, read_only=True)
    ws = wb["All Jobs"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    jobs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            jobs.append(dict(zip(headers, row)))
    wb.close()
    return jsonify(jobs)


@app.route("/download/excel")
def download_excel():
    path = str(OUTPUT_DIR / "jobs.xlsx")
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return "Not found", 404


@app.route("/download/file/<path:filename>")
def download_file(filename):
    return send_from_directory(str(TAILORED_DIR), filename, as_attachment=True)


@app.route("/download/all")
def download_all():
    buf = BytesIO()
    output_str = str(OUTPUT_DIR)
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(output_str):
            for f in files:
                filepath = os.path.join(root, f)
                arcname = os.path.relpath(filepath, output_str)
                zf.write(filepath, arcname)
    buf.seek(0)
    return send_file(buf, mimetype="application/zip", as_attachment=True, download_name="ai-job-pipeline-results.zip")


@app.route("/status")
def status():
    return jsonify({"running": _pipeline_running})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    port = 5000
    print(f"\n  AI Job Pipeline — Web GUI")
    print(f"  Open: http://localhost:{port}\n")
    webbrowser.open(f"http://localhost:{port}")
    app.run(host="0.0.0.0", port=port, debug=False, threaded=True)
